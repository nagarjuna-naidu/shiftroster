import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from calendar import monthrange


def get_current_sheet_name():
    return "Employee_Master"


def get_month_and_year_from_master_sheet(sheet):
    month = sheet.cell(row=3, column=5).value
    year = sheet.cell(row=3, column=6).value
    return int(month), int(year)


def get_month_name(month):
    month_names = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    if 1 <= month <= 12:
        return month_names[month - 1]
    else:
        raise ValueError(f"Invalid month: {month}")


def read_employee(sheet, row_number):
    employee_name = sheet.cell(row=row_number, column=1).value
    wo_days_value = sheet.cell(row=row_number, column=2).value
    wo_days = [] if wo_days_value is None else list(map(int, wo_days_value.split(",")))
    shift = sheet.cell(row=row_number, column=3).value
    return {"employeeName": employee_name, "shift": shift, "wo": wo_days}


def read_shifts(sheet, header_end_row=3, no_of_employees=13):
    employees = []
    for row in range(header_end_row + 1, header_end_row + 1 + no_of_employees):
        employees.append(read_employee(sheet, row))
    return employees


def generate_shift_roster(employees, month, year):
    days_in_month = get_days_in_month(month, year)
    shift_roster = []

    for employee in employees:
        shifts = []
        wo_days = generate_wo_days(employee)
        current_shift = employee['shift']

        for day in range(1, days_in_month + 1):
            if day in wo_days:
                shifts.append("WO")
            else:
                shifts.append(current_shift)

        shift_roster.append({"employee": employee, "nextShifts": shifts})

    return shift_roster


def get_days_in_month(month, year):
    return monthrange(year, month)[1]


def generate_wo_days(employee):
    return employee['wo']


def write_shifts(emp_shifts, sheet, month, year):
    start_row = 1
    generate_month_header(sheet, start_row, month, year)

    shift_colors = {
        'S1': 'FFFF00',  # yellow
        'S2': 'ADD8E6',  # light blue
        'S3': '90EE90',  # light green
        'WO': '808080'   # grey
    }

    # Sort employees by shift: S1 first, then S2, then S3
    emp_shifts.sort(key=lambda x: (x['employee']['shift'] not in shift_colors.keys(), x['employee']['shift']))

    for i, emp_shift in enumerate(emp_shifts, start=2):
        employee_cell = sheet.cell(row=i, column=1)
        employee_cell.value = emp_shift['employee']['employeeName']
        employee_shift_color = shift_colors.get(emp_shift['employee']['shift'])

        if employee_shift_color:
            employee_cell.fill = PatternFill(start_color=employee_shift_color, end_color=employee_shift_color, fill_type="solid")

        for j, shift in enumerate(emp_shift['nextShifts'], start=2):
            current_cell = sheet.cell(row=i, column=j)
            current_cell.value = shift
            shift_color = shift_colors.get(shift)
            if shift_color:
                current_cell.fill = PatternFill(start_color=shift_color, end_color=shift_color, fill_type="solid")


def generate_month_header(sheet, start_row, month, year):
    days_in_month = get_days_in_month(month, year)
    day_headers = [f"{day + 1}-{get_month_name(month)}" for day in range(days_in_month)]

    for i, header in enumerate(day_headers, start=2):
        cell = sheet.cell(row=start_row, column=i)
        cell.value = header
        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')


def main(file_path):
    wb = openpyxl.load_workbook(file_path)
    current_sheet_name = get_current_sheet_name()
    current_sheet = wb[current_sheet_name]

    month, year = get_month_and_year_from_master_sheet(current_sheet)
    new_sheet_name = f"{get_month_name(month)}-{year}"

    if new_sheet_name not in wb.sheetnames:
        new_sheet = wb.create_sheet(new_sheet_name)
    else:
        new_sheet = wb[new_sheet_name]

    employees = read_shifts(current_sheet)
    shift_roster = generate_shift_roster(employees, month, year)
    write_shifts(shift_roster, new_sheet, month, year)

    # Save to a new Excel file
    new_file_path = r"C:\Users\U6048730\Downloads\Shiftroaster\Shift_Roster.xlsx"  # Change to your desired path
    wb.save(new_file_path)
    print(f"Shift roster saved to {new_file_path}")


if __name__ == "__main__":
    main(r"C:\Users\U6048730\Downloads\Shiftroaster\Employee_Master.xlsx")
